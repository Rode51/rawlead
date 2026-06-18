"""Import TG join queue from export txt (wave 3) or owner xlsx (wave 5)."""

from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass
from pathlib import Path
from urllib.parse import urlparse, urlunparse

_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_ROOT / "src"))

from tg_join_lib import read_lines, read_queue_csv, write_queue_csv  # noqa: E402

EXPORT = _ROOT / "docs" / "ops" / "TG_CHANNELS_EXPORT.txt"
BLOCKLIST = _ROOT / "docs" / "ops" / "TG_JOIN_BLOCKLIST.txt"
DROPLIST = _ROOT / "docs" / "ops" / "TG_DROPLIST_2026-05-26.txt"
QUEUE = _ROOT / "docs" / "ops" / "TG_JOIN_QUEUE.csv"
QUEUE_V4 = _ROOT / "docs" / "ops" / "TG_JOIN_QUEUE_v4.csv"
CHATS_XLSX = _ROOT / "docs" / "ops" / "import" / "TG_FREELANCE_CHATS_1500.xlsx"
CHANNELS_XLSX = _ROOT / "docs" / "ops" / "import" / "TG_FREELANCE_CHANNELS_2000.xlsx"
EXISTING_QUEUES = (
    QUEUE,
    _ROOT / "docs" / "ops" / "TG_JOIN_QUEUE_v2.csv",
    _ROOT / "docs" / "ops" / "TG_JOIN_QUEUE_v3.csv",
)

ACCOUNTS = ("acc1", "acc2", "acc3")
WAVE = "3"
NOTES = "волна3"
WAVE_V4 = "5"
NOTES_V4 = "owner-2026-06-15"
FIELDNAMES_V4 = [
    "wave",
    "account",
    "status",
    "name",
    "link",
    "chat_id",
    "notes",
    "kind",
    "subscribers",
]


@dataclass(frozen=True)
class XlsxEntry:
    name: str
    link: str
    subscribers: int
    kind: str


def normalize_link(link: str) -> str:
    s = link.strip()
    if not s:
        return ""
    if s.startswith("t.me/"):
        s = "https://" + s
    elif s.startswith("http://"):
        s = "https://" + s[7:]
    parsed = urlparse(s)
    host = (parsed.hostname or "").lower()
    if host in ("t.me", "telegram.me"):
        host = "t.me"
    path = parsed.path.rstrip("/")
    if not path:
        path = "/"
    return urlunparse(("https", host, path, "", parsed.query, ""))


def link_display_name(link: str, fallback: str = "") -> str:
    path = urlparse(link).path.strip("/")
    if path.startswith("+"):
        return f"invite:{path[1:8]}…" if len(path) > 9 else f"invite:{path[1:]}"
    slug = path.split("/")[-1] or fallback or link
    return slug


def load_blocklist(path: Path) -> set[str]:
    return {normalize_link(ln) for ln in read_lines(path) if normalize_link(ln)}


def load_export(path: Path) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for ln in read_lines(path):
        norm = normalize_link(ln)
        if not norm or norm in seen:
            continue
        seen.add(norm)
        out.append(norm)
    return out


def account_at(index: int) -> str:
    return ACCOUNTS[index % len(ACCOUNTS)]


def load_existing_by_link(queue_paths: tuple[Path, ...] | list[Path]) -> dict[str, dict[str, str]]:
    existing: dict[str, dict[str, str]] = {}
    for path in queue_paths:
        if not path.is_file():
            continue
        _, rows = read_queue_csv(path)
        for row in rows:
            norm = normalize_link(row.get("link", ""))
            if norm:
                existing[norm] = row
    return existing


def read_xlsx_entries(path: Path, *, kind: str) -> list[XlsxEntry]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit("openpyxl required: pip install openpyxl") from exc

    if not path.is_file():
        raise SystemExit(f"Missing xlsx: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        out: list[XlsxEntry] = []
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row or len(row) < 3:
                continue
            name = str(row[0] or "").strip()
            subs_raw = row[1]
            link_raw = row[2]
            if link_raw is None:
                continue
            link = normalize_link(str(link_raw))
            if not link:
                continue
            try:
                subscribers = int(float(subs_raw))
            except (TypeError, ValueError):
                continue
            out.append(
                XlsxEntry(
                    name=name,
                    link=link,
                    subscribers=subscribers,
                    kind=kind,
                )
            )
        return out
    finally:
        wb.close()


def merge_xlsx_sources(
    chats_path: Path,
    channels_path: Path,
) -> tuple[list[XlsxEntry], int]:
    """Return unique entries (first wins) and raw row count before dedupe."""
    raw_count = 0
    seen: set[str] = set()
    merged: list[XlsxEntry] = []
    for path, kind in ((chats_path, "chat"), (channels_path, "channel")):
        entries = read_xlsx_entries(path, kind=kind)
        raw_count += len(entries)
        for entry in entries:
            if entry.link in seen:
                continue
            seen.add(entry.link)
            merged.append(entry)
    return merged, raw_count


def filter_xlsx_entries(
    entries: list[XlsxEntry],
    *,
    min_subscribers: int,
    existing_by_link: dict[str, dict[str, str]],
    blocklist: set[str],
    droplist: set[str],
) -> tuple[list[XlsxEntry], dict[str, int]]:
    stats = {
        "source_unique": len(entries),
        "skip_subscribers": 0,
        "skip_done": 0,
        "skip_duplicate": 0,
        "skip_blocklist": 0,
        "skip_droplist": 0,
        "added": 0,
        "by_acc": {a: 0 for a in ACCOUNTS},
    }
    accepted: list[XlsxEntry] = []
    for entry in entries:
        if entry.subscribers <= min_subscribers:
            stats["skip_subscribers"] += 1
            continue
        if entry.link in blocklist:
            stats["skip_blocklist"] += 1
            continue
        if entry.link in droplist:
            stats["skip_droplist"] += 1
            continue
        if entry.link in existing_by_link:
            status = existing_by_link[entry.link].get("status", "").strip().lower()
            if status == "done":
                stats["skip_done"] += 1
            else:
                stats["skip_duplicate"] += 1
            continue
        accepted.append(entry)
        stats["added"] += 1
    return accepted, stats


def build_v4_rows(
    entries: list[XlsxEntry],
    *,
    pending_before: int,
) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for index, entry in enumerate(entries):
        acc = account_at(pending_before + index)
        rows.append(
            {
                "wave": WAVE_V4,
                "account": acc,
                "status": "pending",
                "name": link_display_name(entry.link, entry.name),
                "link": entry.link,
                "chat_id": "",
                "notes": NOTES_V4,
                "kind": entry.kind,
                "subscribers": str(entry.subscribers),
            }
        )
    return rows


def print_xlsx_stats(stats: dict[str, int], *, dry_run: bool, pending_before: int) -> None:
    mode = "DRY-RUN" if dry_run else "WRITE"
    print(f"[{mode}] xlsx unique links: {stats['source_unique']}")
    print(f"  skip subscribers: {stats['skip_subscribers']}")
    print(f"  добавлено pending: {stats['added']}")
    print(f"  пропуск done:       {stats['skip_done']}")
    print(f"  пропуск дубль:      {stats['skip_duplicate']}")
    print(f"  пропуск blocklist:  {stats['skip_blocklist']}")
    print(f"  пропуск droplist:   {stats['skip_droplist']}")
    for acc in ACCOUNTS:
        count = sum(1 for _i in range(stats["added"]) if account_at(pending_before + _i) == acc)
        print(f"  новых на {acc}:      {count}")
    pending_after = pending_before + stats["added"]
    print(f"  pending всего:      {pending_after} (было {pending_before})")


def run_xlsx_import(*, dry_run: bool, min_subscribers: int) -> int:
    entries, _raw_rows = merge_xlsx_sources(CHATS_XLSX, CHANNELS_XLSX)
    blocklist = load_blocklist(BLOCKLIST)
    droplist = load_blocklist(DROPLIST)
    existing_by_link = load_existing_by_link(EXISTING_QUEUES)

    accepted, stats = filter_xlsx_entries(
        entries,
        min_subscribers=min_subscribers,
        existing_by_link=existing_by_link,
        blocklist=blocklist,
        droplist=droplist,
    )

    fieldnames, existing_rows = read_queue_csv(QUEUE_V4)
    if not fieldnames:
        fieldnames = list(FIELDNAMES_V4)
    pending_before = sum(
        1 for row in existing_rows if row.get("status", "").strip().lower() == "pending"
    )
    new_rows = build_v4_rows(accepted, pending_before=pending_before)
    for index, row in enumerate(new_rows):
        acc = account_at(pending_before + index)
        stats["by_acc"][acc] = stats["by_acc"].get(acc, 0) + 1

    print_xlsx_stats(stats, dry_run=dry_run, pending_before=pending_before)

    if not dry_run and new_rows:
        write_queue_csv(fieldnames, existing_rows + new_rows, QUEUE_V4)
        print(f"Записано: {QUEUE_V4}")
    elif not dry_run:
        print("CSV без изменений.")
    return 0


def run_import(*, dry_run: bool) -> int:
    if not EXPORT.is_file():
        print(f"Нет файла: {EXPORT}")
        return 1

    blocklist = load_blocklist(BLOCKLIST)
    export_links = load_export(EXPORT)
    fieldnames, rows = read_queue_csv(QUEUE)

    if not fieldnames:
        fieldnames = [
            "wave",
            "account",
            "status",
            "name",
            "link",
            "chat_id",
            "notes",
        ]

    existing_by_link: dict[str, dict[str, str]] = {}
    for row in rows:
        norm = normalize_link(row.get("link", ""))
        if norm:
            existing_by_link[norm] = row

    stats = {
        "added": 0,
        "skip_done": 0,
        "skip_duplicate": 0,
        "skip_blocklist": 0,
        "by_acc": {a: 0 for a in ACCOUNTS},
    }
    new_rows: list[dict[str, str]] = []
    pending_before = sum(
        1 for r in rows if r.get("status", "").strip().lower() == "pending"
    )
    assign_index = pending_before

    for link in export_links:
        if link in blocklist:
            stats["skip_blocklist"] += 1
            continue
        if link in existing_by_link:
            status = existing_by_link[link].get("status", "").strip().lower()
            if status == "done":
                stats["skip_done"] += 1
            else:
                stats["skip_duplicate"] += 1
            continue

        acc = account_at(assign_index)
        assign_index += 1
        new_rows.append(
            {
                "wave": WAVE,
                "account": acc,
                "status": "pending",
                "name": link_display_name(link),
                "link": link,
                "chat_id": "",
                "notes": NOTES,
            }
        )
        stats["added"] += 1
        stats["by_acc"][acc] += 1

    out_rows = rows + new_rows

    mode = "DRY-RUN" if dry_run else "WRITE"
    print(f"[{mode}] EXPORT: {len(export_links)} уникальных ссылок")
    print(f"  добавлено pending: {stats['added']}")
    print(f"  пропуск done:       {stats['skip_done']}")
    print(f"  пропуск дубль:      {stats['skip_duplicate']}")
    print(f"  пропуск blocklist:  {stats['skip_blocklist']}")
    for acc in ACCOUNTS:
        print(f"  новых на {acc}:      {stats['by_acc'][acc]}")
    pending_after = pending_before + stats["added"]
    print(f"  pending всего:      {pending_after} (было {pending_before})")

    if not dry_run and new_rows:
        write_queue_csv(fieldnames, out_rows, QUEUE)
        print(f"Записано: {QUEUE}")
    elif not dry_run:
        print("CSV без изменений.")

    return 0


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Import TG join queue from export txt or owner xlsx",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Только счётчики, без записи CSV",
    )
    parser.add_argument(
        "--from-xlsx",
        action="store_true",
        help="Wave 5: TG_FREELANCE_* xlsx → TG_JOIN_QUEUE_v4.csv",
    )
    parser.add_argument(
        "--min-subscribers",
        type=int,
        default=5000,
        help="Wave 5 filter: col B Участники must be > this value (default 5000)",
    )
    args = parser.parse_args()
    if args.from_xlsx:
        raise SystemExit(run_xlsx_import(dry_run=args.dry_run, min_subscribers=args.min_subscribers))
    raise SystemExit(run_import(dry_run=args.dry_run))


if __name__ == "__main__":
    main()
